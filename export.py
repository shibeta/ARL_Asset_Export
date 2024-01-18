import requests
# ARL灯塔的ssl证书为自颁发，关闭requests的InsecureRequestWarning
from urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 代理设置
proxies = {'http': None, 'https': None, 'smtp': None}
# proxies = {'http': "http://127.0.0.1:8080", 'https': "http://127.0.0.1:8080", 'smtp': None}

# 目标网页的URL
url = 'https://<your_ip_here>:5003'

# api key信息
token = '<your_ARL_token_here>'

# 资产范围id
scope_id = '<your_scopeid_here>'

session = requests.Session()
headers = {
  'token': token
}
session.headers.update(headers)

asset_site = [['站点URL', '域名', 'ip', '标题', '状态码', '响应头', '指纹']]  # 站点
asset_domain = [['域名', '解析类型', '记录值']]  # 域名
asset_ip = [['ip', '开放端口', '服务识别', '关联域名', '操作系统', 'Geo_Country', 'Geo_Region', 'Geo_city', 'ASN', 'AS']]  # ip


def get_asset_site(scope_id:str):
  '''
  获取资产站点信息
  scope_id: 资产范围ID
  '''
  # 获取站点数量
  response = session.get(url=url+'/api/asset_site/', params={'scope_id':scope_id, 'size':1}, verify=False, proxies=proxies)

  # 检查请求是否成功
  if response.status_code == 200:
    # print(response.text)
    # 使用json解析响应体
    total = int(response.json()['total'])
  else:
    print(f'请求失败，状态码: {response.status_code}')
    return -1
  
  if total <= 0:
    print('资产范围内没有站点')
    return 0
  print(f"获取到{total}条站点")
  
  for i in range(0, total // 100 + 1):
    response = session.get(url=url+'/api/asset_site/', params={'scope_id':scope_id, 'size':100, 'page':i+1}, verify=False, proxies=proxies)

    if response.status_code == 200:
      json_data = response.json()
      for site in json_data['items']:
        site_info = []
        site_info.append(site['site'])
        site_info.append(site['hostname'])
        site_info.append(site['ip'])
        site_info.append(site['title'])
        site_info.append(site['status'])
        site_info.append(site['headers'])

        site_finger = []
        for finger in site['finger']:
          site_finger.append(finger['name'])

        site_info.append('\n'.join(site_finger))
        asset_site.append(site_info)
    else:
      print(f"请求失败，状态码: {response.status_code}")
      return -1
  return 0


def get_asset_domain(scope_id:str):
  '''
  获取资产域名信息
  scope_id: 资产范围ID
  '''
  # 获取域名数量
  response = session.get(url=url+'/api/asset_domain/', params={'scope_id':scope_id, 'size':1}, verify=False, proxies=proxies)

  # 检查请求是否成功
  if response.status_code == 200:
    # print(response.text)
    # 使用json解析响应体
    total = int(response.json()['total'])
  else:
    print(f'请求失败，状态码: {response.status_code}')
    return -1
  
  if total <= 0:
    print('资产范围内没有域名')
    return 0
  print(f"获取到{total}条域名")
  
  for i in range(0, total // 100 + 1):
    response = session.get(url=url+'/api/asset_domain/', params={'scope_id':scope_id, 'size':100, 'page':i+1}, verify=False, proxies=proxies)

    if response.status_code == 200:
      json_data = response.json()
      for domain in json_data['items']:
        domain_info = []
        domain_info.append(domain['domain'])
        domain_info.append(domain['type'])
        domain_info.append('\n'.join(domain['record']))
        asset_domain.append(domain_info)
    else:
      print(f"请求失败，状态码: {response.status_code}")
      return -1
  return 0


def get_asset_ip(scope_id:str):
  '''
  获取资产IP信息
  scope_id: 资产范围ID
  '''
  # 获取IP数量
  response = session.get(url=url+'/api/asset_ip/', params={'scope_id':scope_id, 'size':1}, verify=False, proxies=proxies)

  # 检查请求是否成功
  if response.status_code == 200:
    # print(response.text)
    # 使用json解析响应体
    total = int(response.json()['total'])
  else:
    print(f'请求失败，状态码: {response.status_code}')
    return -1
  
  if total <= 0:
    print('资产范围内没有IP')
    return 0
  print(f"获取到{total}条IP")
  
  for i in range(0, total // 100 + 1):
    response = session.get(url=url+'/api/asset_ip/', params={'scope_id':scope_id, 'size':100, 'page':i+1}, verify=False, proxies=proxies)

    if response.status_code == 200:
      json_data = response.json()
      for ip in json_data['items']:
        ip_info = []
        ip_info.append(ip['ip'])

        ip_port = []
        ip_port_service = []
        for port in ip['port_info']:
          ip_port.append(port['port_id'])
          ip_port_service.append((port['port_id'], port['service_name']))
        ip_info.append('\n'.join(str(_) for _ in ip_port))
        ip_info.append('\n'.join(["{}:{}".format(*_) for _ in ip_port_service]))

        ip_info.append('\n'.join(ip['domain']))
        # 没有os指纹时, os_info为空字典
        if 'name' in ip['os_info']:
          ip_info.append(ip['os_info']['name'])
        else:
          ip_info.append('')
        # 没有geo信息时, geo_xity为空字典
        if 'country_name' in ip['geo_city']:
          ip_info.append(ip['geo_city']['country_name'])
          ip_info.append(ip['geo_city']['region_name'])
          ip_info.append(ip['geo_city']['city'])
        else:
          [ip_info.append('') for _ in range(3)]
        # 没有asn信息时, geo_asn为空字典
        if 'number' in ip['geo_asn']:
          ip_info.append(ip['geo_asn']['number'])
          ip_info.append(ip['geo_asn']['organization'])
        else:
          [ip_info.append('') for _ in range(2)]
        asset_ip.append(ip_info)
    else:
      print(f"请求失败，状态码: {response.status_code}")
      return -1
  return 0


def save_to_xlsx(filename:str):
  workbook = Workbook()

  sheet = workbook.active
  sheet.title = '站点'
  for row in asset_site:
    sheet.append(row)
  for row_index, row in enumerate(sheet.iter_rows(min_row=2, min_col=6, max_col=7), start=2):
    for cell in row:
      cell.alignment = Alignment(wrap_text=True)
  
  sheet = workbook.create_sheet(title='域名')
  for row in asset_domain:
    sheet.append(row)
  for row_index, row in enumerate(sheet.iter_rows(min_row=2, min_col=3, max_col=3), start=2):
    for cell in row:
      cell.alignment = Alignment(wrap_text=True)

  sheet = workbook.create_sheet(title='IP')
  for row in asset_ip:
    sheet.append(row)
  for row_index, row in enumerate(sheet.iter_rows(min_row=2, min_col=2, max_col=4), start=2):
    for cell in row:
      cell.alignment = Alignment(wrap_text=True)
  
  workbook.save(filename)

if __name__ == "__main__":
  if get_asset_site(scope_id) < 0:
    exit()
  if get_asset_domain(scope_id) < 0:
    exit()
  if get_asset_ip(scope_id) < 0:
    exit()

  save_to_xlsx(f'{scope_id}.xlsx')
  print(f'结果保存到{scope_id}.xlsx')
